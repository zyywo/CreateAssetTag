from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import load_workbook
import datetime
from tag import Tag
pdfmetrics.registerFont(TTFont('Vera', 'simsun.ttc'))


class AssetTags():
    def __init__(self, excel_file, index_of_header, out_pdf_file):
        self.__header, self.__datas = self.__loadfile(excel_file, index_of_header)
        self.tag = Tag(out_pdf_file)

    @staticmethod
    def __loadfile(excel_, header_at_row):
        """参数：excel文件；表头所在的行数
           返回值：表头内容（列表）；excel内容（字典）
        """
        wb = load_workbook(excel_, read_only=True)
        st = wb.active
        header = []
        data={}

        for i in range(1, st.max_column + 1):
            header.append(st.cell(row=header_at_row, column=i).value)

        for i, row in enumerate(st.rows):
            temp_dict = {}
            for j,cell in enumerate(row):
                if isinstance(cell.value,datetime.datetime):
                    temp_dict[header[j]] = str(cell.value)[:-8]
                elif cell.value is None:
                    temp_dict[header[j]] = ''
                else:
                    temp_dict[header[j]] = cell.value
            data[i] = temp_dict
        return header, data

    def create_page(self, header=None, _fontsize=7):
        if header is None: header=self.__header

        for i,d in enumerate(self.__datas):
            #因为是字典，所以d的值为与i的值一样

            self.tag.addpage(header,self.__datas[i],_fontsize)

    def save(self):
        self.tag.build()

if __name__ == '__main__':

    my_custom = ['资产名称','购置日期','资产品牌','规格型号','序列号','MAC地址','所属部门','所在区域']

    in_file = '东莞仓固定资产-测试.xlsx'
    out_file = '测试.pdf'

    pdf = AssetTags(in_file, 2, out_file)

    pdf.create_page(my_custom,7)

    pdf.save()
